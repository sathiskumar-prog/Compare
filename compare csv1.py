import os
import io
import pandas as pd
from flask import Flask, request, render_template_string, send_file, jsonify
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

app = Flask(__name__)

# Global storage for the session
storage = {
    "df1": None, "df2": None, 
    "removed": None, "added": None, "mismatches": None, 
    "orig_cols": [], "names": ["", ""]
}

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Delta Engine Pro | Render</title>
    <style>
        :root { --primary: #4f46e5; --bg: #f8fafc; --border: #e2e8f0; --danger: #ef4444; --success: #10b981; --warning: #f59e0b; }
        body { font-family: 'Inter', system-ui, sans-serif; background: var(--bg); color: #1e293b; margin: 0; padding: 20px; }
        .container { max-width: 1200px; margin: auto; }
        .card { background: white; padding: 2rem; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.05); border: 1px solid var(--border); margin-bottom: 2rem; }
        
        /* Layout Utilities */
        .flex-between { display: flex; justify-content: space-between; align-items: center; width: 100%; }
        .hidden { display: none !important; }
        
        /* File Upload */
        .file-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; }
        .drop-zone { border: 2px dashed #cbd5e1; padding: 30px; text-align: center; border-radius: 8px; background: #fafafa; position: relative; cursor: pointer; }
        .drop-zone input { position: absolute; inset: 0; opacity: 0; cursor: pointer; width: 100%; height: 100%; }
        
        button { background: var(--primary); color: white; border: none; padding: 12px 24px; border-radius: 8px; font-weight: 700; cursor: pointer; transition: 0.2s; }
        button:hover { filter: brightness(1.1); }

        /* Results Tables */
        .table-wrap { overflow-x: auto; max-height: 400px; border: 1px solid var(--border); border-radius: 8px; margin-top: 15px; background: white; }
        table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
        th, td { padding: 10px 15px; border: 1px solid var(--border); white-space: nowrap; text-align: left; }
        th { background: #f1f5f9; position: sticky; top: 0; color: #64748b; font-weight: 600; }
        tr:nth-child(even) { background: #f8fafc; }
        
        /* Highlights */
        .mismatch-red { color: var(--danger); font-weight: bold; background: #fff1f2; padding: 2px 4px; border-radius: 4px; }
        .btn-xlsx { background: #10b981; color: white; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-weight: bold; font-size: 0.85rem; }
        
        #loader { display: none; text-align: center; color: var(--primary); font-weight: bold; margin: 10px 0; }
    </style>
</head>
<body>
    <div class="container">
        <h1 style="text-align:center; font-weight:900; color:#111827;">Delta Engine Pro</h1>
        
        <div class="card" id="input-card">
            <div id="upload-stage">
                <div class="file-grid">
                    <div class="drop-zone">
                        <strong>Base File (A)</strong><br><span id="n1" style="font-size:0.8rem; color:var(--primary)">Upload CSV</span>
                        <input type="file" id="f1" accept=".csv" onchange="document.getElementById('n1').innerText=this.files[0].name">
                    </div>
                    <div class="drop-zone">
                        <strong>New File (B)</strong><br><span id="n2" style="font-size:0.8rem; color:var(--primary)">Upload CSV</span>
                        <input type="file" id="f2" accept=".csv" onchange="document.getElementById('n2').innerText=this.files[0].name">
                    </div>
                </div>
                <button onclick="uploadAndAnalyze()" style="width:100%">Analyze Columns</button>
            </div>

            <div id="key-stage" class="hidden">
                <p><strong>Step 2: Select Key Column(s)</strong></p>
                <div id="keys-box" style="display:grid; grid-template-columns:repeat(auto-fill, minmax(160px, 1fr)); gap:10px; margin-bottom:20px; padding:15px; background:#f1f5f9; border-radius:10px;"></div>
                <button onclick="runComparison()" style="width:100%">Run Comparison</button>
            </div>
            <div id="loader">Processing...</div>
        </div>

        <div id="results" class="hidden">
            <div class="card">
                <div class="flex-between" style="margin-bottom: 2rem;">
                    <h2 style="margin:0">Analysis Results</h2>
                    <button onclick="location.reload()" style="background:#64748b; width:auto">Reset All</button>
                </div>

                <div style="margin-bottom:3rem">
                    <div class="flex-between">
                        <h3 style="color:var(--danger); margin:0;">🔴 Removed Rows</h3>
                        <a href="/export/removed" class="btn-xlsx">Download XLSX</a>
                    </div>
                    <div id="out-del" class="table-wrap"></div>
                </div>

                <div style="margin-bottom:3rem">
                    <div class="flex-between">
                        <h3 style="color:var(--success); margin:0;">🟢 Added Rows</h3>
                        <a href="/export/added" class="btn-xlsx">Download XLSX</a>
                    </div>
                    <div id="out-add" class="table-wrap"></div>
                </div>

                <div>
                    <div class="flex-between">
                        <h3 style="color:var(--warning); margin:0;">🟡 Mismatches</h3>
                        <a href="/export/mismatches" class="btn-xlsx">Download XLSX</a>
                    </div>
                    <div id="out-mis" class="table-wrap"></div>
                </div>
            </div>
        </div>
    </div>

    <script>
        async function uploadAndAnalyze() {
            const f1 = document.getElementById('f1').files[0];
            const f2 = document.getElementById('f2').files[0];
            if(!f1 || !f2) return alert("Select both files.");

            document.getElementById('loader').style.display = 'block';
            const fd = new FormData();
            fd.append('file_a', f1);
            fd.append('file_b', f2);

            const res = await fetch('/analyze', { method: 'POST', body: fd });
            const data = await res.json();
            
            document.getElementById('keys-box').innerHTML = data.columns.map(c => `
                <label style="font-size:0.85rem; cursor:pointer"><input type="checkbox" value="${c}" checked> ${c}</label>
            `).join('');
            
            document.getElementById('upload-stage').classList.add('hidden');
            document.getElementById('key-stage').classList.remove('hidden');
            document.getElementById('loader').style.display = 'none';
        }

        async function runComparison() {
            const keys = Array.from(document.querySelectorAll('#keys-box input:checked')).map(i => i.value);
            if(keys.length === 0) return alert("Select a key.");

            document.getElementById('loader').style.display = 'block';
            const res = await fetch('/compare', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ keys })
            });
            const data = await res.json();

            document.getElementById('out-del').innerHTML = data.del_h;
            document.getElementById('out-add').innerHTML = data.add_h;
            document.getElementById('out-mis').innerHTML = data.mis_h;
            
            document.getElementById('results').classList.remove('hidden');
            document.getElementById('input-card').classList.add('hidden');
            document.getElementById('loader').style.display = 'none';
        }
    </script>
</body>
</html>
"""

def apply_excel_styles(df, type_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        ws = writer.sheets['Results']
        
        # Formatting setup
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        red_font = Font(color="FF0000", bold=True)

        for cell in ws[1]:
            cell.fill, cell.border, cell.font = header_fill, border, Font(bold=True)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = border
                if row_idx % 2 == 0: cell.fill = alt_fill
                if type_name == "mismatches" and " → " in str(cell.value):
                    cell.font = red_font
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22
    output.seek(0)
    return output

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/analyze', methods=['POST'])
def analyze():
    f1, f2 = request.files['file_a'], request.files['file_b']
    df1 = pd.read_csv(f1).fillna('').astype(str).apply(lambda x: x.str.strip())
    df2 = pd.read_csv(f2).fillna('').astype(str).apply(lambda x: x.str.strip())
    storage.update({"df1": df1, "df2": df2, "orig_cols": list(df1.columns)})
    return jsonify({"columns": [c for c in df1.columns if c in df2.columns]})

@app.route('/compare', methods=['POST'])
def compare():
    keys = request.json['keys']
    df1, df2, orig_cols = storage['df1'], storage['df2'], storage['orig_cols']
    merged = df1.merge(df2, on=keys, how='outer', indicator=True, suffixes=('_old', '_new'))
    
    def safe_order(df, order): return df[[c for c in order if c in df.columns]]

    res_del = safe_order(merged[merged['_merge'] == 'left_only'], orig_cols)
    res_add = merged[merged['_merge'] == 'right_only'].rename(columns={c+'_new': c for c in df2.columns if c not in keys})
    res_add = safe_order(res_add, list(df2.columns))
    
    both = merged[merged['_merge'] == 'both']
    others = [c for c in orig_cols if c not in keys]
    mis_ui, mis_raw = [], []
    for _, r in both.iterrows():
        row_ui, row_raw, diff = {}, {}, False
        for c in others:
            v1, v2 = r.get(c+'_old', ''), r.get(c+'_new', '')
            if v1 != v2:
                row_ui[c] = f'<span class="mismatch-red">{v1} → {v2}</span>'
                row_raw[c] = f"{v1} → {v2}"
                diff = True
            else: row_ui[c] = row_raw[c] = v1
        if diff:
            for k in keys: row_ui[k] = row_raw[k] = r[k]
            mis_ui.append(row_ui); mis_raw.append(row_raw)
    
    res_mis_df = pd.DataFrame(mis_raw)
    if not res_mis_df.empty: res_mis_df = safe_order(res_mis_df, orig_cols)
    storage.update({"removed": res_del, "added": res_add, "mismatches": res_mis_df})

    return jsonify({
        "del_h": res_del.to_html(index=False),
        "add_h": res_add.to_html(index=False),
        "mis_h": safe_order(pd.DataFrame(mis_ui), orig_cols).to_html(index=False, escape=False) if mis_ui else "No mismatches."
    })

@app.route('/export/<type>')
def export(type):
    df = storage.get(type if type != "mismatches" else "mismatches")
    if df is None or (isinstance(df, pd.DataFrame) and df.empty): return "No data", 404
    return send_file(apply_excel_styles(df, type), as_attachment=True, download_name=f"delta_{type}.xlsx")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 5000)))