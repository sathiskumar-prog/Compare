import os
import io
import pandas as pd
from flask import Flask, request, render_template_string, send_file
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

app = Flask(__name__)

# Global storage (Clears on Render restart/idle)
storage = {
    "df1": None, "df2": None, 
    "removed": None, "added": None, "mismatches": None, 
    "orig_cols": []
}

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Delta Engine | Order-Locked</title>
    <style>
        :root { --primary: #4f46e5; --bg: #f8fafc; --border: #e2e8f0; }
        body { font-family: system-ui, sans-serif; background: var(--bg); color: #1e293b; padding: 20px; }
        .container { max-width: 1200px; margin: auto; }
        .card { background: white; padding: 2rem; border-radius: 12px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); border: 1px solid var(--border); margin-bottom: 2rem; }
        .file-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; }
        .drop-zone { border: 2px dashed #cbd5e1; padding: 25px; text-align: center; border-radius: 8px; background: #fafafa; cursor: pointer; position: relative; }
        .drop-zone input { position: absolute; inset: 0; opacity: 0; cursor: pointer; }
        button { background: var(--primary); color: white; border: none; padding: 12px; border-radius: 8px; font-weight: 700; cursor: pointer; width: 100%; }
        .table-wrap { overflow-x: auto; max-height: 400px; border: 1px solid var(--border); margin-top: 15px; background: white; }
        table { width: 100%; border-collapse: collapse; font-size: 0.85rem; border: 1px solid var(--border); }
        th, td { padding: 10px; border: 1px solid var(--border); white-space: nowrap; text-align: left; }
        th { background: #f1f5f9; position: sticky; top: 0; }
        tr:nth-child(even) { background: #f8fafc; }
        .mismatch-red { color: #ef4444; font-weight: bold; background: #fff1f2; padding: 2px 4px; border-radius: 3px; }
        .btn-xlsx { background: #10b981; color: white; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-weight: bold; font-size: 0.8rem; display: inline-block; }
    </style>
</head>
<body>
    <div class="container">
        <h2 style="font-weight:900;">Delta Engine <small style="font-weight:400; color:#64748b;">| Ordered Output</small></h2>
        <div class="card">
            <form method="POST" enctype="multipart/form-data">
                {% if not cols %}
                <div class="file-grid">
                    <div class="drop-zone"><strong>Base File (A)</strong><br><input type="file" name="file1" accept=".csv" required></div>
                    <div class="drop-zone"><strong>New File (B)</strong><br><input type="file" name="file2" accept=".csv" required></div>
                </div>
                <button type="submit" name="action" value="upload">Analyze Structure</button>
                {% else %}
                <div style="margin-bottom:1rem"><strong>Select Anchor Keys:</strong></div>
                <div style="display:grid; grid-template-columns:repeat(auto-fill, minmax(150px, 1fr)); gap:10px; margin-bottom:20px">
                    {% for col in cols %}
                    <label style="font-size: 0.85rem;"><input type="checkbox" name="keys" value="{{ col }}" checked> {{ col }}</label>
                    {% endfor %}
                </div>
                <button type="submit" name="action" value="compare">Run Ordered Comparison</button>
                {% endif %}
            </form>
        </div>

        {% if results %}
        <div class="card">
            <h3>🔴 Removed Rows <a href="/export/removed" class="btn-xlsx">Export XLSX</a></h3>
            <div class="table-wrap">{{ results.del_h|safe }}</div>
            <h3 style="margin-top:40px">🟢 Added Rows <a href="/export/added" class="btn-xlsx">Export XLSX</a></h3>
            <div class="table-wrap">{{ results.add_h|safe }}</div>
            <h3 style="margin-top:40px">🟡 Mismatches <a href="/export/mismatches" class="btn-xlsx">Export XLSX</a></h3>
            <div class="table-wrap">{{ results.mis_h|safe }}</div>
        </div>
        {% endif %}
    </div>
</body>
</html>
"""

def get_ordered_df(df, order_list):
    """Filters only columns that exist and sorts them per the original list."""
    existing = [c for c in order_list if c in df.columns]
    return df[existing]

def style_excel(df, type_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Delta_Results')
        ws = writer.sheets['Delta_Results']

        # Styling
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
        red_font = Font(color="FF0000", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.border = border
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = border
                if row_idx % 2 == 0: cell.fill = alt_fill
                if type_name == "mismatches" and " >> " in str(cell.value):
                    cell.font = red_font
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    output.seek(0)
    return output

@app.route('/', methods=['GET', 'POST'])
def index():
    cols, results = None, None
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'upload':
            f1, f2 = request.files['file1'], request.files['file2']
            df1 = pd.read_csv(f1).fillna('').astype(str).apply(lambda x: x.str.strip())
            df2 = pd.read_csv(f2).fillna('').astype(str).apply(lambda x: x.str.strip())
            # Save the Original Column Blueprint
            storage.update({"df1": df1, "df2": df2, "orig_cols": list(df1.columns)})
            cols = [c for c in df1.columns if c in df2.columns]
        
        elif action == 'compare':
            keys = request.form.getlist('keys')
            df1, df2, orig_cols = storage['df1'], storage['df2'], storage['orig_cols']
            
            merged = df1.merge(df2, on=keys, how='outer', indicator=True, suffixes=('_old', '_new'))
            
            # 1. Removed (Force File A Order)
            res_del = get_ordered_df(merged[merged['_merge'] == 'left_only'], orig_cols)
            
            # 2. Added (Force File B Order)
            res_add = merged[merged['_merge'] == 'right_only'].rename(columns={c+'_new': c for c in df2.columns if c not in keys})
            res_add = get_ordered_df(res_add, list(df2.columns))
            
            # 3. Mismatches (Force File A Order)
            both = merged[merged['_merge'] == 'both']
            others = [c for c in orig_cols if c not in keys]
            mis_ui, mis_raw = [], []
            for _, r in both.iterrows():
                row_ui, row_raw, diff = {}, {}, False
                for c in others:
                    v1, v2 = r.get(c+'_old', ''), r.get(c+'_new', '')
                    if v1 != v2:
                        row_ui[c] = f'<span class="mismatch-red">{v1} → {v2}</span>'
                        row_raw[c] = f"{v1} >> {v2}"
                        diff = True
                    else: row_ui[c] = row_raw[c] = v1
                if diff:
                    for k in keys: row_ui[k] = row_raw[k] = r[k]
                    mis_ui.append(row_ui); mis_raw.append(row_raw)

            storage.update({"removed": res_del, "added": res_add, "mismatches": pd.DataFrame(mis_raw)})
            
            # Final Check: Reorder Mismatches per orig_cols before rendering
            mis_final_ui = pd.DataFrame(mis_ui)
            if not mis_final_ui.empty:
                mis_final_ui = get_ordered_df(mis_final_ui, orig_cols)

            results = {
                "del_h": res_del.to_html(index=False),
                "add_h": res_add.to_html(index=False),
                "mis_h": mis_final_ui.to_html(index=False, escape=False) if not mis_final_ui.empty else "No mismatches found."
            }
    return render_template_string(HTML_TEMPLATE, cols=cols, results=results)

@app.route('/export/<type>')
def export(type):
    df = storage.get(type if type != "mismatches" else "mismatches")
    if df is None or (isinstance(df, pd.DataFrame) and df.empty): return "No data", 404
    
    # Final safety: Ensure export df follows the correct column order
    if type == "removed" or type == "mismatches":
        df = get_ordered_df(df, storage['orig_cols'])
        
    styled_file = style_excel(df, type)
    return send_file(styled_file, as_attachment=True, download_name=f"delta_{type}.xlsx")

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)